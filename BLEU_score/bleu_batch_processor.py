import pandas as pd
import sys
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

# 导入包含BLEU评分函数的模块
# 假设这两个文件分别是 汉语处理demo 和 英文处理demo，请根据实际文件名进行调整
from 汉语处理demo import *
from 英文处理demo import *


def process_translations_file(input_file, output_file):
    """
    处理输入文件中的翻译内容，并将BLEU评分结果输出到一个Excel文件中，
    中文和英文结果分别放在不同的工作表，严格按照指定的列顺序排列。
    增加了在输出内容最后一行添加平均值计算处理，计算每列的平均值，并在第一列写入"平均得分："。
    另外，设置每个平均 combined 得分的单元格背景颜色为 RGB(141, 180, 226).

    注意：现在输入Excel文件第一行为标题行（无关数据），实际数据从第二行开始。
    读取顺序为：
      - 中文：B列为中文翻译参考，后面依次为：ours, 网易, 百度, Google, deepl, 腾讯翻译君
      - 英文：I列为英文参考，后面依次为：ours, 网易, 百度, Google, deepl, 腾讯翻译君
    输出到Excel时，将从第二列开始输出，第一列保留，用于显示"平均得分："
    """
    try:
        # 读取Excel文件，跳过第一行标题行
        print(f"正在读取文件: {input_file}")
        df = pd.read_excel(input_file, header=None, skiprows=1)
        print(f"成功读取文件，共 {len(df)} 行")

        # 初始化中文和英文结果列表
        chinese_results = []
        english_results = []

        # 定义列索引：
        # 中文部分：B列（索引1）为中文参考，
        # 接下来依次为：ours (索引2), 网易 (索引3), 百度 (索引4), Google (索引5), deepl (索引6), 腾讯翻译君 (索引7)
        chinese_ref_col = 1  # 中文参考所在B列

        # 英文部分：I列（索引8）为英文参考，
        # 接下来依次为：ours (索引9), 网易 (索引10), 百度 (索引11), Google (索引12), deepl (索引13), 腾讯翻译君 (索引14)
        english_ref_col = 8  # 英文参考所在I列

        print("开始处理翻译评分...")
        processed_rows = 0
        skipped_rows = 0
        error_rows = 0

        for idx in range(len(df)):
            try:
                row = df.iloc[idx]

                # 处理中文部分：
                chinese_ref = str(row.iloc[chinese_ref_col])
                # 按照读取顺序：ours, 网易, 百度, Google, deepl, 腾讯翻译君
                chinese_ours = str(row.iloc[chinese_ref_col + 1])
                chinese_netease = str(row.iloc[chinese_ref_col + 2])
                chinese_baidu = str(row.iloc[chinese_ref_col + 3])
                chinese_google = str(row.iloc[chinese_ref_col + 4])
                chinese_deepl = str(row.iloc[chinese_ref_col + 5])
                chinese_tencent = str(row.iloc[chinese_ref_col + 6])

                # 处理英文部分：
                english_ref = str(row.iloc[english_ref_col])
                english_ours = str(row.iloc[english_ref_col + 1])
                english_netease = str(row.iloc[english_ref_col + 2])
                english_baidu = str(row.iloc[english_ref_col + 3])
                english_google = str(row.iloc[english_ref_col + 4])
                english_deepl = str(row.iloc[english_ref_col + 5])
                english_tencent = str(row.iloc[english_ref_col + 6])

                # 跳过任何无效或空的参考
                if pd.isna(chinese_ref) or pd.isna(english_ref) or chinese_ref == "nan" or english_ref == "nan":
                    skipped_rows += 1
                    continue

                # 计算中文BLEU分数 - 使用导入的函数
                chinese_ours_scores = calculate_chinese_bleu_complete(chinese_ref, chinese_ours)
                chinese_netease_scores = calculate_chinese_bleu_complete(chinese_ref, chinese_netease)
                chinese_baidu_scores = calculate_chinese_bleu_complete(chinese_ref, chinese_baidu)
                chinese_google_scores = calculate_chinese_bleu_complete(chinese_ref, chinese_google)
                chinese_deepl_scores = calculate_chinese_bleu_complete(chinese_ref, chinese_deepl)
                chinese_tencent_scores = calculate_chinese_bleu_complete(chinese_ref, chinese_tencent)

                # 计算英文BLEU分数 - 使用导入的函数
                english_ours_scores = calculate_english_bleu_complete(english_ref, english_ours)
                english_netease_scores = calculate_english_bleu_complete(english_ref, english_netease)
                english_baidu_scores = calculate_english_bleu_complete(english_ref, english_baidu)
                english_google_scores = calculate_english_bleu_complete(english_ref, english_google)
                english_deepl_scores = calculate_english_bleu_complete(english_ref, english_deepl)
                english_tencent_scores = calculate_english_bleu_complete(english_ref, english_tencent)

                # 输出结果按照读取顺序： ours, 网易, 百度, Google, deepl, 腾讯翻译君
                chinese_row = [
                    # ours
                    chinese_ours_scores['unigram_bleu'],
                    chinese_ours_scores['bigram_bleu'],
                    chinese_ours_scores['combined_bleu'],
                    # 网易
                    chinese_netease_scores['unigram_bleu'],
                    chinese_netease_scores['bigram_bleu'],
                    chinese_netease_scores['combined_bleu'],
                    # 百度
                    chinese_baidu_scores['unigram_bleu'],
                    chinese_baidu_scores['bigram_bleu'],
                    chinese_baidu_scores['combined_bleu'],
                    # Google
                    chinese_google_scores['unigram_bleu'],
                    chinese_google_scores['bigram_bleu'],
                    chinese_google_scores['combined_bleu'],
                    # deepl
                    chinese_deepl_scores['unigram_bleu'],
                    chinese_deepl_scores['bigram_bleu'],
                    chinese_deepl_scores['combined_bleu'],
                    # 腾讯翻译君
                    chinese_tencent_scores['unigram_bleu'],
                    chinese_tencent_scores['bigram_bleu'],
                    chinese_tencent_scores['combined_bleu']
                ]
                chinese_results.append(chinese_row)

                english_row = [
                    # ours
                    english_ours_scores['unigram_bleu'],
                    english_ours_scores['bigram_bleu'],
                    english_ours_scores['combined_bleu'],
                    # 网易
                    english_netease_scores['unigram_bleu'],
                    english_netease_scores['bigram_bleu'],
                    english_netease_scores['combined_bleu'],
                    # 百度
                    english_baidu_scores['unigram_bleu'],
                    english_baidu_scores['bigram_bleu'],
                    english_baidu_scores['combined_bleu'],
                    # Google
                    english_google_scores['unigram_bleu'],
                    english_google_scores['bigram_bleu'],
                    english_google_scores['combined_bleu'],
                    # deepl
                    english_deepl_scores['unigram_bleu'],
                    english_deepl_scores['bigram_bleu'],
                    english_deepl_scores['combined_bleu'],
                    # 腾讯翻译君
                    english_tencent_scores['unigram_bleu'],
                    english_tencent_scores['bigram_bleu'],
                    english_tencent_scores['combined_bleu']
                ]
                english_results.append(english_row)

                processed_rows += 1
                if processed_rows % 10 == 0:
                    print(f"已处理 {processed_rows} 行...")
            except Exception as e:
                print(f"处理第 {idx + 1} 行时出错: {e}")
                error_rows += 1
                continue

        print(f"处理完成。成功处理 {processed_rows} 行，跳过 {skipped_rows} 行，错误 {error_rows} 行。")

        # 定义新的列名顺序，共18列，按照顺序： ours, 网易, 百度, Google, deepl, 腾讯翻译君
        original_column_names = [
            'ours_1-gram', 'ours_2-gram', 'ours_combined',
            '网易_1-gram', '网易_2-gram', '网易_combined',
            '百度_1-gram', '百度_2-gram', '百度_combined',
            'Google_1-gram', 'Google_2-gram', 'Google_combined',
            'deepl_1-gram', 'deepl_2-gram', 'deepl_combined',
            '腾讯翻译君_1-gram', '腾讯翻译君_2-gram', '腾讯翻译君_combined'
        ]
        # 将数据转换为DataFrame
        chinese_df = pd.DataFrame(chinese_results, columns=original_column_names)
        english_df = pd.DataFrame(english_results, columns=original_column_names)

        # 添加平均值行到中文和英文评分表（原来平均行仅包含数值）
        chinese_avg_row = chinese_df.mean(numeric_only=True).to_dict()
        english_avg_row = english_df.mean(numeric_only=True).to_dict()
        # 追加平均行
        chinese_df = pd.concat([chinese_df, pd.DataFrame([chinese_avg_row])], ignore_index=True)
        english_df = pd.concat([english_df, pd.DataFrame([english_avg_row])], ignore_index=True)

        # 插入空白的第一列，使输出从第二列开始
        chinese_df.insert(0, "标签", "")
        english_df.insert(0, "标签", "")

        # 在最后一行第一列写入"平均得分："
        chinese_df.at[chinese_df.index[-1], "标签"] = "平均得分："
        english_df.at[english_df.index[-1], "标签"] = "平均得分："

        # 保存到Excel文件并设置所有列宽为15.88
        print(f"正在保存结果到 {output_file}...")
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            chinese_df.to_excel(writer, sheet_name="中文BLEU评分", index=False)
            english_df.to_excel(writer, sheet_name="英文BLEU评分", index=False)

            # 添加元数据表
            metadata = pd.DataFrame({
                "处理日期": [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                "输入文件": [input_file],
                "处理的中文句子数": [len(chinese_results)],
                "处理的英文句子数": [len(english_results)]
            })
            metadata.to_excel(writer, sheet_name="元数据", index=False)

            # 定义平均行中需要设置背景色的“combined”列（基于最终工作表的列字母）
            # 根据DataFrame列，索引（0-indexed）映射到Excel列字母:
            # 插入的第一列 -> A, then:
            # ours_combined: DataFrame列索引 3 -> Excel列 D
            # 网易_combined: index 6 -> Excel列 G
            # 百度_combined: index 9 -> Excel列 J
            # Google_combined: index 12 -> Excel列 M
            # deepl_combined: index 15 -> Excel列 P
            # 腾讯翻译君_combined: index 18 -> Excel列 S
            combined_columns = ['D', 'G', 'J', 'M', 'P', 'S']
            fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")

            # 设置所有工作表的列宽为15.88，并为平均行的combined单元格添加背景颜色
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for col in worksheet.columns:
                    col_letter = col[0].column_letter
                    worksheet.column_dimensions[col_letter].width = 15.88

                # 针对非元数据工作表，设置平均行（最后一行）中指定列的背景颜色
                if sheet_name != "元数据":
                    max_row = worksheet.max_row  # 平均行所在行号
                    for col in combined_columns:
                        cell = worksheet[f"{col}{max_row}"]
                        cell.fill = fill

        print(f"处理完成。结果已保存到 {output_file}")

    except Exception as e:
        print(f"处理文件时出错: {e}")
        raise


if __name__ == "__main__":
    # if len(sys.argv) < 3:
    #     print("用法: python bleu_batch_processor_import.py <输入文件.xlsx> <输出文件.xlsx>")
    #     print("注意: 输出文件应该使用.xlsx扩展名")
    #     print("注意: 请确保汉语处理demo和英文处理demo文件在当前目录中")
    #     sys.exit(1)
    input_file = "all_translation_results2.xlsx"
    output_file = "score_output2.xlsx"

    if not output_file.endswith(".xlsx"):
        output_file += ".xlsx"

    print(f"批处理开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"输入文件: {input_file}")
    print(f"输出文件: {output_file}")

    process_translations_file(input_file, output_file)

    print(f"批处理结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")